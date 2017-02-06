# -*- coding: utf-8 -*- 

import os
import openpyxl
import ctypes
import json
import word2vec
import pymongo

WIKIPEDIA_PATH="../data/wikipedia/boxed_content_20160407/"
PREPROCESSED_STORY_PATH="../data/wikipedia/preprocessed/"
MORP_LEMMA_ONLY_STORY_PATH="../data/wikipedia/morp_lemma_only/"
MORP_LEMMA_TYPE_STORY_PATH="../data/wikipedia/morp_lemma_type/"
MORP_LEMMA_ONLY_QUERY_PATH="../data/query/morp_lemma_only/"
MORP_LEMMA_TYPE_QUERY_PATH="../data/query/morp_lemma_type/"
QUERY_PATH="../data/query/"
QUERY_FILE_LIST=["001.도전과제_단문QA.xlsx","002.일반분야_단문QA.xlsx"]

def all_word_to_one(document_path,output_path):
    file_list=os.listdir(document_path)
    with open(output_path,"a") as out:
        for f_name in file_list:
            if f_name.endswith(".txt")==True:
                out.write(open(document_path+f_name).read()+"\n")


def vectorize_sentence(sentence_list,vocabulary,maxlen):
    vectorized_sentence_list=[]
    for sentence in sentence_list:
        if maxlen!=None:
            vectorized_sentence=[]
            for word in sentence:
                if word not in vocabulary:
                    vectorized_sentence.append(len(vocabulary))
                else:
                    vectorized_sentence.append(vocabulary.index(word)+1)
            if len(vectorized_sentence)<maxlen:
                vectorized_sentence+=[0]*(maxlen-len(vectorized_sentence))
        else:
            vectorized_sentence=[0]*len(vocabulary)
            for word in sentence:
                if word not in vocabulary:
                    vectorized_sentence[-1]+=1
                else:
                    vectorized_sentence[vocabulary.index(word)]+=1
        vectorized_sentence_list.append(vectorized_sentence)
    return vectorized_sentence_list

def load_stories_morp():
    story_content=[]
    story_file_list=os.listdir(MORP_LEMMA_TYPE_STORY_PATH)
    for story_file in story_file_list:
        for line in open(MORP_LEMMA_TYPE_STORY_PATH+story_file).read().split("\n"):
            if line.split(" ")[:-1]!=[]:
                #story_content.append(line.split(" ")[:-1])
                story_content+=line.split(" ")[:-1]
                if story_content[-1]!=".":
                    story_content+="."
    return story_content

def load_queries_morp():
    query_content=[]
    query_file_list=os.listdir(MORP_LEMMA_TYPE_QUERY_PATH)
    for query_file in query_file_list:
        if query_file!="001_multiple_query.out":
            continue
        query_start=False
        for line in open(MORP_LEMMA_TYPE_QUERY_PATH+query_file).read().split("\n"):
            if line.split(" ")[0]=="":
                break
            elif line.split(" ")[0][0]=="#":
                query=line.split(" ")[1:-1]
                query_start=True
            elif query_start==True:
                answer=line.split(" ")[:-2]
                query_start=False
                if len(answer)==1:
                    query_content.append([query,answer])
    return query_content

def load_our_queries_morp():
    query_content=[]
    query_file_list=os.listdir(MORP_LEMMA_TYPE_QUERY_PATH)
    for query_file in query_file_list:
        if query_file!="our_simple_query.out":
            continue
        query_start=False
        idx=1
        for line in open(MORP_LEMMA_TYPE_QUERY_PATH+query_file).read().split("\n"):
            if line.split(" ")[0]=="":
                break
            elif line.split(" ")[0][0]=="#":
                query=line.split(" ")[1:-1]
                query_start=True
            elif query_start==True:
                answer=line.split(" ")[:-2]
                query_start=False
                if len(answer)==1:
                    query_content.append([query,answer,idx])
                if len(answer)==2:
                    if answer[1] in ['개/NNB','년/NNB','명/NNB','강/NNG','색/NNG','쌍/NNG','번째/NNB','항/NNG','분/NNB','위/NNB','메달/NNG','대/NNB']:
                        query_content.append([query,[answer[0]],idx])
                idx+=1
    return query_content

def load_stories():
    story_content_list=[]
    story_file_list=os.listdir(MORP_LEMMA_ONLY_STORY_PATH)
    for story_file in story_file_list:
        for line in open(MORP_LEMMA_ONLY_STORY_PATH+story_file).read().split("\n"):
            if line.split(" ")[:-1]!=[]:
                story_content=[]
                #story_content.append(line.split(" ")[:-1])
                story_content+=line.split(" ")[:-1]
                if story_content[-1]!=".":
                    story_content+="."
                story_content_list.append(story_content)
    return story_content_list

def load_pymongo():
    return pymongo.MongoClient("localhost",27017)

def load_stories_with_query_using_title_more(query):
    search_word_list=[]
    for word in query:
        if "NN" in word.split("/")[1]:
            search_word_list.append(word)
    search_word_list=list(set(search_word_list))

    search_file_list=[]
    for search_word in search_word_list:
        os.system("csearch -l \""+search_word+"\" > ./csearch_out/load_stories_with_query_using_title_out.txt")
        for line in open("./csearch_out/load_stories_with_query_using_title_out.txt"):
            if line!="":
                search_file_list.append(line)
    search_file_list=list(set(search_file_list))

    story_content_list=[]
    for search_file in search_file_list:
        title_flag=True
        for line in open(search_file[:-1]).read().split("\n"):
            if title_flag==True:
                for search_word in search_word_list:
                    if search_word in line.split(" ")[:-1]:
                        title_flag=False
                        break
            if title_flag==True:
                break
            if line.split(" ")[:-1]!=[]:
                for search_word in search_word_list:
                    if search_word in line.split(" ")[:-1]:
                        story_content=[]
                        #story_content.append(line.split(" ")[:-1])
                        story_content+=line.split(" ")[:-1]
                        if story_content[-1]!=".":
                            story_content+="."
                        story_content_list.append(story_content)
                        break
    return story_content_list

def load_stories_with_query_using_title(query):
    search_word_list=[]
    for word in query:
        if "NN" in word.split("/")[1]:
            search_word_list.append(word)
    search_word_list=list(set(search_word_list))

    search_file_list=[]
    for search_word in search_word_list:
        os.system("csearch -l \""+search_word+"\" > ./csearch_out/load_stories_with_query_using_title_out.txt")
        for line in open("./csearch_out/load_stories_with_query_using_title_out.txt"):
            if line!="":
                search_file_list.append(line)
    search_file_list=list(set(search_file_list))

    story_content_list=[]
    for search_file in search_file_list:
        title_flag=True
        for line in open(search_file[:-1]).read().split("\n"):
            if title_flag==True:
                for search_word in search_word_list:
                    if search_word in line.split(" ")[:-1]:
                        title_flag=False
                        break
            if title_flag==True:
                break
            if line.split(" ")[:-1]!=[]:
                for search_word in search_word_list:
                    if search_word in line.split(" ")[:-1]:
                        story_content=[]
                        #story_content.append(line.split(" ")[:-1])
                        story_content+=line.split(" ")[:-1]
                        if story_content[-1]!=".":
                            story_content+="."
                        story_content_list.append(story_content)
                        break
    return story_content_list

def load_stories_with_query(query):
    search_word_list=[]
    for word in query:
        if "NN" in word.split("/")[1]:
            search_word_list.append(word)
    search_word_list=list(set(search_word_list))

    search_file_list=[]
    for search_word in search_word_list:
        os.system("csearch -l \""+search_word+"\" > ./csearch_out/load_stories_with_query_out.txt")
        for line in open("./csearch_out/load_stories_with_query_out.txt"):
            if line!="":
                search_file_list.append(line)
    search_file_list=list(set(search_file_list))

    story_content_list=[]
    for search_file in search_file_list:
        for line in open(search_file[:-1]).read().split("\n"):
            if line.split(" ")[:-1]!=[]:
                for search_word in search_word_list:
                    if search_word in line.split(" ")[:-1]:
                        story_content=[]
                        #story_content.append(line.split(" ")[:-1])
                        story_content+=line.split(" ")[:-1]
                        if story_content[-1]!=".":
                            story_content+="."
                        story_content_list.append(story_content)
                        break
    return story_content_list

def load_stories_with_our_query_from_filename(search_file):
    story_content_list=[]
    for line in open(MORP_LEMMA_ONLY_STORY_PATH+search_file).read().split("\n"):
        if line.split(" ")[:-1]!=[]:
            story_content=[]
            #story_content.append(line.split(" ")[:-1])
            story_content+=line.split(" ")[:-1]
            if story_content[-1]!=".":
                story_content+="."
            story_content_list.append(story_content)
    return story_content_list


def load_stories_with_our_query(query):
    query_id=query[-1]
    wb=openpyxl.load_workbook(QUERY_PATH+"한국어_단문_QA.xlsx",read_only=True)
    sheet=wb[wb.get_sheet_names()[0]]
    query_list=list(sheet.iter_rows())
    search_file=None

    for query_elem in query_list:
        if query_elem[0].value==None:
            break
        if query_elem[0].value==str(query_id):
            search_file=query_elem[3].value
            break
    story_content_list=[]
    for line in open(MORP_LEMMA_TYPE_STORY_PATH+search_file).read().split("\n"):
        if line.split(" ")[:-1]!=[]:
            story_content=[]
            #story_content.append(line.split(" ")[:-1])
            story_content+=line.split(" ")[:-1]
            if story_content[-1]!=".":
                story_content+="."
            story_content_list.append(story_content)
    return story_content_list

def load_stories_supporting_fact_with_our_query(query):
    query_id=query[-1]
    wb=openpyxl.load_workbook(QUERY_PATH+"한국어_단문_QA.xlsx",data_only=True)
    sheet=wb[wb.get_sheet_names()[0]]
    query_list=list(sheet.iter_rows())
    search_file=None
    supporting_fact_list=None
    for query_elem in query_list:
        if query_elem[0].value==None:
            break
        if query_elem[0].value==str(query_id):
            search_file=query_elem[3].value
            supporting_fact_list=list(map(int,query_elem[4].value.split(",")))
            break
    story_content_list=[]
    for line in open(MORP_LEMMA_TYPE_STORY_PATH+search_file).read().split("\n"):
        if line.split(" ")[:-1]!=[]:
            story_content=[]
            #story_content.append(line.split(" ")[:-1])
            story_content+=line.split(" ")[:-1]
            if story_content[-1]!=".":
                story_content+="."
            story_content_list.append(story_content)
    return story_content_list,supporting_fact_list
    

def load_queries():
    query_content=[]
    query_file_list=os.listdir(MORP_LEMMA_ONLY_QUERY_PATH)
    for query_file in query_file_list:
        if query_file!="001_multiple_query.out":
            continue
        query_start=False
        for line in open(MORP_LEMMA_ONLY_QUERY_PATH+query_file).read().split("\n"):
            if line.split(" ")[0]=="#":
                query=line.split(" ")[1:-1]
                query_start=True
            elif query_start==True:
                answer=line.split(" ")[:-2]
                query_start=False
                if len(answer)==1:
                    query_content.append([query,answer])
    return query_content



def preprocess_preprocessed_data(file_name,PREPROCESSED_PATH):
    content=""
    with open(PREPROCESSED_PATH+file_name) as f:
        content=f.read().split("The IP address is checked")[1].split("----끝----")[0]
    return content

def extract_morp_our(file_name,PREPROCESSED_PATH,MORP_LEMMA_ONLY_PATH,MORP_LEMMA_TYPE_PATH):
    first_flag=True
    try:
        json_content_list=json.load(open(PREPROCESSED_PATH+file_name))
    except ValueError:
        json_content_list=json.loads(preprocess_preprocessed_data(file_name,PREPROCESSED_PATH))
    num_query=int(len(json_content_list)/2)

    for idx in range(num_query):
        for sentence in json_content_list[idx*2]['sentence'][1:]:
            for morp_word in sentence['morp']:
                if first_flag==True:
                    with open(MORP_LEMMA_ONLY_PATH+file_name,"w") as f:
                        f.write(morp_word['lemma']+" ")
                    with open(MORP_LEMMA_TYPE_PATH+file_name,"w") as f:
                        f.write(morp_word['lemma']+"/"+morp_word['type']+" ")
                    first_flag=False
                else:
                    with open(MORP_LEMMA_ONLY_PATH+file_name,"a") as f:
                        f.write(morp_word['lemma']+" ")
                    with open(MORP_LEMMA_TYPE_PATH+file_name,"a") as f:
                        f.write(morp_word['lemma']+"/"+morp_word['type']+" ")
            with open(MORP_LEMMA_ONLY_PATH+file_name,"a") as f:
                f.write("\n")
            with open(MORP_LEMMA_TYPE_PATH+file_name,"a") as f:
                f.write("\n")
        for sentence in json_content_list[idx*2+1]['sentence'][1:]:
            for morp_word in sentence['morp']:
                with open(MORP_LEMMA_ONLY_PATH+file_name,"a") as f:
                    f.write(morp_word['lemma']+" ")
                with open(MORP_LEMMA_TYPE_PATH+file_name,"a") as f:
                    f.write(morp_word['lemma']+"/"+morp_word['type']+" ")
            with open(MORP_LEMMA_ONLY_PATH+file_name,"a") as f:
                f.write("\n")
            with open(MORP_LEMMA_TYPE_PATH+file_name,"a") as f:
                f.write("\n")
    return file_name

def extract_morp(file_name,PREPROCESSED_PATH,MORP_LEMMA_ONLY_PATH,MORP_LEMMA_TYPE_PATH):
    first_flag=True
    try:
        json_content=json.load(open(PREPROCESSED_PATH+file_name))
    except ValueError:
        json_content=json.loads(preprocess_preprocessed_data(file_name,PREPROCESSED_PATH))
    for sentence in json_content['sentence']:
        for morp_word in sentence['morp']:
            if first_flag==True:
                with open(MORP_LEMMA_ONLY_PATH+file_name,"w") as f:
                    f.write(morp_word['lemma']+" ")
                with open(MORP_LEMMA_TYPE_PATH+file_name,"w") as f:
                    f.write(morp_word['lemma']+"/"+morp_word['type']+" ")
                first_flag=False
            else:
                with open(MORP_LEMMA_ONLY_PATH+file_name,"a") as f:
                    f.write(morp_word['lemma']+" ")
                with open(MORP_LEMMA_TYPE_PATH+file_name,"a") as f:
                    f.write(morp_word['lemma']+"/"+morp_word['type']+" ")
        with open(MORP_LEMMA_ONLY_PATH+file_name,"a") as f:
            f.write("\n")
        with open(MORP_LEMMA_TYPE_PATH+file_name,"a") as f:
            f.write("\n")
    return file_name

def load_query(multiple_choice_only=True):
    for query_file in QUERY_FILE_LIST:
        if multiple_choice_only==True:
            if query_file!="001.도전과제_단문QA.xlsx":
                continue
        wb=openpyxl.load_workbook(QUERY_PATH+query_file,read_only=True)
        sheet=wb[wb.get_sheet_names()[0]]
        query_list=list(sheet.iter_rows())
        column_name=[]
        for elem in query_list[0]:
            column_name.append(elem.value)
        query_list=query_list[1:]
        for query in query_list:
            if multiple_choice_only==True:
                if query[3].value==None:
                    continue
            print([elem.value for elem in query])
            input()
    

def read_file(file_name):
    content=""
    with open(WIKIPEDIA_PATH+file_name) as f:
        while(True):
            line=f.readline()
            if line=="":
                break
            content+=line
    return content

def process_content(content):
    title=None
    text=""
    categori=None
    page_struct=None
    text_start=False
    text_list=False
    for line in content.split("\n"):
        if line[:7]=="<TITLE>":
            if title==None:
                title=line[7:]
            elif title!=line[7:]:
                print("There is other title??")
                print(title)
                print(line[7:])
                exit()
        elif line[:13]=="<PAGE_STRUCT>":
            page_struct=line.split("__<__")[1].split("__>__")[0]
            if len(page_struct.split("\t"))==2:
                page_struct=page_struct.split("\t")[1]
            text_start=True
        elif ("__<__" in line)&("__>__" in line):
            if text_start==True:
                if line.split("__<__")[1].split("__>__")[0][-1]!=".":
                    text_list=True
                    text+=page_struct+":"
                    text+=line.split("__<__")[1].split("__>__")[0]+", "
                else:
                    text+=line.split("__<__")[1].split("__>__")[0]+" "
                text_start=False
            else:
                if text_list==True:
                    text+=line.split("__<__")[1].split("__>__")[0]+", "
                else:
                    text+=line.split("__<__")[1].split("__>__")[0]+" "
        elif line[:10]=="<CATEGORY>":
            if text_list==True:
                text=text[:-2]+". "
                text_list=False
            if categori==None:
                categori=line[10:].split("\t")
            elif categori!=line[10:].split("\t"):
                print("There is other categori??")
                print(categori)
                print(line[:10])
                exit()
    text+="분류: "+", ".join(categori)
    return title, text
